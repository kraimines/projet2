import requests
import json
import re
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from PIL import Image
from pdf2image import convert_from_path
import pytesseract
import docling
from pathlib import Path
from django.shortcuts import render
from .forms import TicketUploadForm
from .models import ExtractionHistory, TicketHistory, AccountingEntry
from doctr.models import ocr_predictor
from doctr.io import DocumentFile
from docling.document_converter import DocumentConverter
import os
import logging
from django.conf import settings
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
from decimal import Decimal
from datetime import datetime, date
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def diagnose_system():
    """Diagnostic function to check system health"""
    issues = []
    
    # Check Ollama connection
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code != 200:
            issues.append("Ollama server not responding properly")
        else:
            logger.info("✓ Ollama server is running")
    except requests.exceptions.ConnectionError:
        issues.append("Cannot connect to Ollama server - is it running?")
    except requests.exceptions.Timeout:
        issues.append("Ollama server timeout - server may be overloaded")
    except Exception as e:
        issues.append(f"Ollama connection error: {str(e)}")
    
    # Check media directory
    try:
        media_root = getattr(settings, 'MEDIA_ROOT', None)
        if not media_root:
            issues.append("MEDIA_ROOT not configured in settings")
        elif not os.path.exists(media_root):
            issues.append(f"Media directory doesn't exist: {media_root}")
        else:
            logger.info(f"✓ Media directory exists: {media_root}")
    except Exception as e:
        issues.append(f"Media directory check failed: {str(e)}")
    
    # Check required Python packages
    required_packages = ['doctr', 'pytesseract', 'docling', 'PIL', 'pdf2image']
    for package in required_packages:
        try:
            __import__(package)
            logger.info(f"✓ {package} is available")
        except ImportError:
            issues.append(f"Missing required package: {package}")
    
    if issues:
        logger.error("System issues found:")
        for issue in issues:
            logger.error(f"  - {issue}")
    else:
        logger.info("✓ All system checks passed")
    
    return issues

def clean_json_response(text):
    """
    Nettoie une réponse JSON malformée de manière plus robuste
    """
    if not text or not isinstance(text, str):
        return None
    
    print(f"Texte original reçu: {text[:200]}...")
    
    # Supprimer les balises <think> et leur contenu
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    text = re.sub(r'<.*?>', '', text)  # Supprimer toutes les balises HTML/XML
    
    # Nettoyer les espaces et sauts de ligne excessifs
    text = re.sub(r'\n+', '\n', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    # Chercher le JSON le plus externe
    start_positions = [i for i, char in enumerate(text) if char == '{']
    end_positions = [i for i, char in enumerate(text) if char == '}']
    
    if start_positions and end_positions:
        # Essayer le JSON le plus externe d'abord
        outer_start = start_positions[0]
        outer_end = end_positions[-1]
        if outer_end > outer_start:
            json_candidate = text[outer_start:outer_end+1]
            try:
                # Nettoyer le JSON
                cleaned_json = json_candidate.strip()
                # Remplacer les guillemets simples par des guillemets doubles
                cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                # Supprimer les virgules trailing
                cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                
                result = json.loads(cleaned_json)
                # Vérifier que c'est un objet avec les clés principales
                main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                    print(f"✅ JSON externe valide trouvé")
                    return result
            except Exception as e:
                print(f"❌ Erreur parsing JSON externe: {e}")
    
    # Essayer tous les JSON possibles
    best_json = None
    best_score = 0
    
    for start in start_positions:
        for end in end_positions:
            if end > start:
                json_candidate = text[start:end+1]
                try:
                    cleaned_json = json_candidate.strip()
                    cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                    cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                    
                    result = json.loads(cleaned_json)
                    main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                    
                    if isinstance(result, dict):
                        score = sum(1 for k in main_keys if k in result)
                        if score >= 2 and score > best_score:
                            best_json = result
                            best_score = score
                            print(f"✅ JSON trouvé avec score {score}")
                except Exception as e:
                    print(f"❌ Erreur parsing JSON candidat: {e}")
                    continue
    
    if best_json:
        return best_json
    
    # Dernier essai : chercher le JSON avec une regex plus permissive
    # Chercher le dernier JSON valide dans le texte
    json_pattern = r'\{(?:[^{}]|(?:\{[^{}]*\}))*\}'
    json_matches = re.findall(json_pattern, text, re.DOTALL)
    
    # Essayer le dernier JSON trouvé (le plus probable d'être complet)
    for json_match in reversed(json_matches):
        try:
            cleaned_json = json_match.strip()
            cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
            cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
            
            result = json.loads(cleaned_json)
            main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
            
            if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                print(f"✅ JSON trouvé via regex (dernier)")
                return result
        except Exception as e:
            print(f"❌ Erreur parsing JSON regex: {e}")
            continue
    
    # Si rien trouvé, essayer de chercher juste après "}" pour voir s'il y a du JSON valide
    if '}' in text:
        last_brace_pos = text.rfind('}')
        if last_brace_pos > 0:
            # Chercher le début du JSON en remontant
            brace_count = 0
            start_pos = last_brace_pos
            for i in range(last_brace_pos, -1, -1):
                if text[i] == '}':
                    brace_count += 1
                elif text[i] == '{':
                    brace_count -= 1
                    if brace_count == 0:
                        start_pos = i
                        break
            
            if start_pos < last_brace_pos:
                json_candidate = text[start_pos:last_brace_pos + 1]
                try:
                    cleaned_json = json_candidate.strip()
                    cleaned_json = re.sub(r"'([^']*)'", r'"\1"', cleaned_json)
                    cleaned_json = re.sub(r',(\s*[}\]])', r'\1', cleaned_json)
                    
                    result = json.loads(cleaned_json)
                    main_keys = ['Magasin', 'Date', 'NumeroTicket', 'Articles', 'Total']
                    
                    if isinstance(result, dict) and sum(1 for k in main_keys if k in result) >= 2:
                        print(f"✅ JSON trouvé via recherche de braces")
                        return result
                except Exception as e:
                    print(f"❌ Erreur parsing JSON via braces: {e}")
    
    print("❌ Aucun JSON valide trouvé")
    return None



def analyze_three_texts_with_llm_fast(ocr_results):
    """
    Version rapide avec un modèle plus léger
    """
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")
    
    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici trois extraits OCR du même ticket :

--- OCR DocLing ---
{docling_text}

--- OCR Tesseract ---
{tesseract_text}

--- OCR Doctr ---
{doctr_text}

Extrais les éléments suivants et retourne UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "Numéro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÈGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit être inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplémentaire
4. Commence par {{ et termine par }}
"""
    try:
        print("Tentative avec modèle rapide (mistral)...")
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "mistral", "prompt": prompt, "stream": False},
            timeout=30
        )
        print("Réponse reçue de l'API Ollama (modèle rapide)")
        result_text = response.json().get("response", "")
        
        # Parser le JSON de la réponse LLM
        parsed_data = clean_json_response(result_text)
        if parsed_data:
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "Données extraites par modèle rapide",
                "texte_fusionne": result_text.strip()
            }
            
            # Post-traitement pour s'assurer que le timbre fiscal est bien détecté
            result_data = post_process_timbre_fiscal(result_data)
            
            # Validation et correction avec regex
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)
            
            return result_data
        
        # Si pas de JSON, retourner le texte brut
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Texte fusionné et corrigé (modèle rapide)",
            "texte_fusionne": result_text.strip()
        }
    except requests.exceptions.Timeout:
        print("Timeout avec mistral, essai avec modèle ultra-rapide...")
        return analyze_three_texts_with_llm_ultra_fast(ocr_results)
    except Exception as e:
        print("Erreur avec modèle rapide:", str(e))
        return {"error": f"Erreur API (modèle rapide) : {str(e)}"}

def analyze_three_texts_with_llm_ultra_fast(ocr_results):
    """
    Version ultra-rapide avec un modèle très léger
    """
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")
    
    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici trois extraits OCR du même ticket :

--- OCR DocLing ---
{docling_text}

--- OCR Tesseract ---
{tesseract_text}

--- OCR Doctr ---
{doctr_text}

Extrais les éléments suivants et retourne UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "Numéro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÈGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit être inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplémentaire
4. Commence par {{ et termine par }}
"""
    try:
        print("Tentative avec modèle ultra-rapide (llama2)...")
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "llama2", "prompt": prompt, "stream": False},
            timeout=15
        )
        print("Réponse reçue de l'API Ollama (modèle ultra-rapide)")
        result_text = response.json().get("response", "")
        
        # Parser le JSON de la réponse LLM
        parsed_data = clean_json_response(result_text)
        if parsed_data:
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "Données extraites par modèle ultra-rapide",
                "texte_fusionne": result_text.strip()
            }
            
            # Post-traitement pour s'assurer que le timbre fiscal est bien détecté
            result_data = post_process_timbre_fiscal(result_data)
            
            # Validation et correction avec regex
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)
            
            return result_data
        
        # Si pas de JSON, retourner le texte brut
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Texte fusionné et corrigé (modèle ultra-rapide)",
            "texte_fusionne": result_text.strip()
        }
    except Exception as e:
        print("Erreur avec modèle ultra-rapide:", str(e))
        return {"error": f"Erreur API (modèle ultra-rapide) : {str(e)}"}

import os
import json
import re
from openai import OpenAI
import google.generativeai as genai
from dotenv import load_dotenv
load_dotenv()

# Initialiser le client OpenAI avec l'URL HuggingFace
try:
    hf_token = os.environ.get("HF_TOKEN", "")
    if not hf_token:
        print("⚠️ Token HuggingFace manquant dans .env")
        client = None
    else:
        client = OpenAI(
            base_url="https://router.huggingface.co/v1",
            api_key=hf_token,
        )
        print("✅ Client HuggingFace initialisé")
except Exception as e:
    print(f"❌ Erreur initialisation client HuggingFace: {e}")
    client = None

def analyze_three_texts_with_llm(ocr_results):
    docling_text = ocr_results.get("docling", "")
    tesseract_text = ocr_results.get("tesseract", "")
    doctr_text = ocr_results.get("doctr", "")

    if not docling_text.strip() and not tesseract_text.strip() and not doctr_text.strip():
        return {
            "Date": "",
            "Magasin": "",
            "NumeroTicket": "",
            "Total": "",
            "Articles": [],
            "Commentaire": "Aucun texte OCR extrait - les textes sont vides"
        }

    prompt = f"""Tu es un assistant expert en analyse de tickets de caisse.

Voici un extrait OCR du ticket de caisse :

--- OCR Doctr ---
{doctr_text}

Ta tâche est d'extraire les éléments suivants et de retourner UNIQUEMENT un objet JSON valide :

{{
  "Magasin": "Nom du magasin",
  "NumeroTicket": "Numéro du ticket",
  "Date": "JJ/MM/AAAA HH:MM",
  "Articles": [
    {{ "nom": "Nom article", "prix": "Prix en DT" }},
    {{ "nom": "TIMBRE FISCAL", "prix": "0.100 DT" }}
  ],
  "Total": "Montant total en DT"
}}

RÈGLES IMPORTANTES :
1. Le timbre fiscal (0.100 DT, 0.200 DT, etc.) doit être inclus dans Articles avec nom "TIMBRE FISCAL"
2. Si tu vois "100 DT", convertis-le en "0.100 DT" pour les timbres fiscaux
3. Retourne UNIQUEMENT le JSON, sans commentaires ni texte supplémentaire
4. Commence par {{ et termine par }}
5. Assure-toi que tous les montants soient des chaînes terminées par 'DT'

NE RENVOIE QUE UN OBJET JSON. PAS DE COMMENTAIRES, PAS DE TEXTE, PAS DE PENSÉES.
Commence DIRECTEMENT par '{' et termine par '}'.
"""

    try:
        # Vérifier si le client est disponible
        if client is None:
            print("❌ Client HuggingFace non disponible, utilisation du fallback Ollama")
            raise Exception("Client HuggingFace non initialisé")
            
        print("🔄 Appel à l'API HuggingFace avec Qwen...")

        completion = client.chat.completions.create(
            model="Qwen/Qwen3-30B-A3B:novita",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            timeout=30  # Timeout de 30 secondes
        )

        result_text = completion.choices[0].message.content.strip()

        # Clean and extract JSON with error handling
        try:
            # Utiliser la fonction de nettoyage améliorée
            parsed_data = clean_json_response(result_text)
            
            if not parsed_data:
                print("❌ Aucun JSON valide trouvé dans la réponse LLM")
                return {
                    "error": "Invalid JSON response from LLM: No valid JSON found",
                    "raw_response": result_text,
                    "Commentaire": "La réponse LLM ne contient pas de JSON valide"
                }
                
            print(f"✅ JSON parsé avec succès")
            
            result_data = {
                "Date": parsed_data.get("Date", ""),
                "Magasin": parsed_data.get("Magasin", ""),
                "NumeroTicket": parsed_data.get("NumeroTicket", ""),
                "Total": parsed_data.get("Total", ""),
                "Articles": parsed_data.get("Articles", []),
                "Commentaire": "Données extraites via HuggingFace Qwen",
                "texte_fusionne": result_text
            }

            # Post-traitement personnalisé
            result_data = post_process_timbre_fiscal(result_data)
            texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
            result_data = valider_et_corriger_avec_regex(result_data, texte_ocr_combined)

            return result_data
            
        except Exception as e:
            logger.error(f"Erreur lors du parsing JSON: {str(e)}")
            return {
                "error": f"Invalid JSON response from LLM: {str(e)}",
                "raw_response": result_text,
                "Commentaire": f"Erreur lors du parsing JSON: {str(e)}"
            }

    except Exception as e:
        print("Erreur lors de l'appel à l'API HuggingFace:", str(e))
        # Fallback vers les modèles Ollama locaux
        try:
            print("Tentative de fallback vers Ollama...")
            return analyze_three_texts_with_llm_fast(ocr_results)
        except Exception as fallback_error:
            print(f"Erreur avec fallback Ollama: {str(fallback_error)}")
            # Dernier recours : analyse simple avec regex
            try:
                texte_ocr_combined = f"{docling_text}\n{tesseract_text}\n{doctr_text}"
                regex_result = extraire_elements_avec_regex(texte_ocr_combined)
                
                return {
                    "Date": regex_result.get("dates_valides", [""])[0] if regex_result.get("dates_valides") else "",
                    "Magasin": "",
                    "NumeroTicket": "",
                    "Total": regex_result.get("total", ""),
                    "Articles": regex_result.get("articles", []),
                    "Commentaire": f"Analyse par regex (fallback) - Erreur API: {str(e)}",
                    "texte_fusionne": texte_ocr_combined,
                    "ValidationRegex": {
                        "total_coherent": regex_result.get("total_coherent", False),
                        "somme_articles": regex_result.get("somme_articles", ""),
                        "total_detecte": regex_result.get("total", "")
                    }
                }
            except Exception as regex_error:
                return {
                    "Date": "",
                    "Magasin": "",
                    "NumeroTicket": "",
                    "Total": "",
                    "Articles": [],
                    "Commentaire": f"Erreur complète - API: {str(e)}, Fallback: {str(fallback_error)}, Regex: {str(regex_error)}",
                    "texte_fusionne": "Erreur de traitement"
                }

def analyze_three_texts_with_gemini(ocr_results):
    """
    Analyse les 3 textes OCR avec Google Generative AI (Gemini)
    """
    try:
        # Configuration de l'API Google Generative AI
        api_key = os.getenv('GOOGLE_API_KEY')
        if not api_key or api_key == 'your_google_api_key_here':
            # Mode démo - retourner des données d'exemple
            logger.warning("Clé API Google Generative AI manquante - Mode démo activé")
            return {
                "error": "Clé API Google Generative AI manquante. Voici un exemple de résultat.",
                "demo_mode": True,
                "Magasin": "DEMO - Magasin Exemple",
                "Date": "2024-01-15",
                "Heure": "14:30",
                "Numero_ticket": "DEMO-001",
                "Articles": [
                    {
                        "nom": "Article démo 1",
                        "quantite": 2,
                        "prix_unitaire": 5.500,
                        "prix_total": 11.000
                    },
                    {
                        "nom": "TIMBRE FISCAL",
                        "quantite": 1,
                        "prix_unitaire": 0.100,
                        "prix_total": 0.100
                    }
                ],
                "Sous_total": 11.000,
                "Remise": 0.000,
                "Timbre_fiscal": 0.100,
                "Total": 11.100,
                "Methode_paiement": "espèces",
                "TVA_details": {
                    "taux_19": 2.090,
                    "taux_13": 0.000,
                    "taux_7": 0.000
                },
                "model_used": "Google Gemini 1.5 Flash (Mode Démo)",
                "analysis_timestamp": datetime.now().isoformat(),
                "raw_response": "Mode démo - Aucune API appelée",
                "instructions": {
                    "title": "Comment obtenir une clé API Google Generative AI :",
                    "steps": [
                        "1. Allez sur https://makersuite.google.com/app/apikey",
                        "2. Connectez-vous avec votre compte Google",
                        "3. Cliquez sur 'Create API Key'",
                        "4. Copiez la clé générée",
                        "5. Remplacez 'your_google_api_key_here' dans le fichier .env",
                        "6. Redémarrez le serveur Django"
                    ]
                }
            }
        
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        # Fusionner les 3 textes OCR
        texte_fusionne = f"""
        === TEXTE DOCTR ===
        {ocr_results.get('doctr', '')}
        
        === TEXTE TESSERACT ===
        {ocr_results.get('tesseract', '')}
        
        === TEXTE DOCLING ===
        {ocr_results.get('docling', '')}
        """
        
        # Prompt simplifié pour correspondre au format Qwen
        prompt = f"""
        Tu es un expert en extraction de données de tickets de caisse. Analyse ces 3 textes OCR d'un même ticket et extrais les informations suivantes au format JSON strict :
        
        {{
            "Magasin": "nom du magasin",
            "Date": "YYYY-MM-DD",
            "NumeroTicket": "numéro du ticket" ou null,
            "Articles": [
                {{
                    "nom": "nom de l'article",
                    "prix": "prix de l'article avec unité (ex: 5.500 DT)"
                }}
            ],
            "Total": "montant total avec unité (ex: 39.500 DT)"
        }}
        
        RÈGLES IMPORTANTES :
        - Utilise les 3 textes pour obtenir la meilleure précision
        - Corrige les erreurs OCR courantes : O→0, I→1, etc.
        - Si une information n'est pas trouvée, utilise null
        - Réponds UNIQUEMENT avec le JSON, sans texte supplémentaire
        
        TEXTES OCR À ANALYSER :
        {texte_fusionne}
        """
        
        logger.info("Envoi de la requête à Google Generative AI...")
        
        # Générer la réponse
        response = model.generate_content(prompt)
        raw_response = response.text
        
        logger.info(f"Réponse brute reçue de Gemini: {raw_response[:200]}...")
        
        # Nettoyer et parser la réponse JSON
        result_data = clean_json_response(raw_response)
        
        if result_data and isinstance(result_data, dict):
            logger.info("Analyse Gemini réussie")
            
            # Ajouter des métadonnées
            result_data['texte_fusionne'] = texte_fusionne
            result_data['raw_response'] = raw_response
            result_data['model_used'] = 'Google Gemini 1.5 Flash'
            result_data['analysis_timestamp'] = datetime.now().isoformat()
            
            # Validation et correction avec regex
            result_data = valider_et_corriger_avec_regex(result_data, texte_fusionne)
            
            return result_data
        else:
            logger.error("Erreur de parsing JSON Gemini: Réponse invalide")
            return {
                "error": "Réponse JSON invalide de Gemini",
                "raw_response": raw_response,
                "parsed_response": result_data
            }
    
    except Exception as e:
        logger.error(f"Erreur lors de l'analyse avec Gemini: {str(e)}")
        return {
            "error": f"Erreur lors de l'analyse Gemini: {str(e)}",
            "raw_response": None
        }

def extraire_elements_avec_regex(texte):
    """
    Extraction précise d'éléments avec regex pour validation/correction des résultats LLM
    """
    if not texte or not isinstance(texte, str):
        return {}
    
    # Regex pour dates (JJ/MM/AAAA ou JJ-MM-AAAA)
    dates = re.findall(r'\b([0-3]?\d)[/-]([0-1]?\d)[/-](\d{4})\b', texte)
    dates_valides = []
    for d, m, y in dates:
        try:
            date_obj = datetime.strptime(f"{d}/{m}/{y}", "%d/%m/%Y")
            dates_valides.append(date_obj.strftime("%d/%m/%Y"))
        except:
            pass
    
    # Regex pour prix au format comme 0.100 DT, 12.50 DT
    prix_pattern = r'\b\d+\.\d{2,3}\s?DT\b'
    prix_trouves = re.findall(prix_pattern, texte)
    
    # Convertir les prix en float pour calcul (supprimer " DT" et convertir)
    def prix_to_float(p):
        try:
            return float(p.replace("DT", "").strip())
        except Exception:
            return 0.0
    
    prix_floats = [prix_to_float(p) for p in prix_trouves]

    # Regex pour détecter timbre fiscal (ex: 0.100 DT, 0.200 DT etc.)
    timbres = [p for p in prix_trouves if p in ["0.100 DT", "0.200 DT", "0.300 DT", "0.400 DT", "0.500 DT"]]
    timbres_float = [prix_to_float(t) for t in timbres]
    
    # Supposons que le total est indiqué par une ligne avec "Total" + montant
    total_match = re.search(r'(Total|TOTAL|total)\s*[:\-]?\s*(\d+\.\d{2,3})\s?DT', texte)
    total = None
    if total_match:
        try:
            total = float(total_match.group(2))
        except Exception:
            total = None
    
    # Extraire les articles, en filtrant les mots interdits (espèce, rendu, reçu, total, etc.)
    # On suppose que chaque ligne contient nom article + prix (ex: Pain 1.200 DT)
    mots_interdits = ['espece', 'rendu', 'recu', 'total', 'remise', 'timbre', 'fiscal', 'taxe', 'stamp','pece']
    
    articles = []
    # Exemple simple : chercher toutes les lignes avec un prix, puis extraire le nom avant
    lignes = texte.split('\n')
    for ligne in lignes:
        prix_match = re.search(prix_pattern, ligne)
        if prix_match:
            prix = prix_match.group()
            # Extraire nom avant prix
            nom = ligne[:prix_match.start()].strip().lower()
            
            # Vérifier si un mot interdit est dans le nom
            if any(mot in nom for mot in mots_interdits):
                continue  # Ignore cette ligne, ce n'est pas un article
            
            # On suppose que le nom est le texte avant le prix
            nom_article = ligne[:prix_match.start()].strip()
            articles.append({"nom": nom_article, "prix": prix})

    # Vérification heuristique : somme des articles == total
    somme_articles = sum([prix_to_float(a["prix"]) for a in articles])
    total_correct = (total is not None and abs(somme_articles - total) < 0.01)  # marge d'erreur 1 centime
    
    resultat = {
        "dates_valides": dates_valides,
        "timbres_fiscaux": timbres,
        "total": f"{total:.3f} DT" if total is not None else "",
        "articles": articles,
        "somme_articles": f"{somme_articles:.3f} DT",
        "total_coherent": total_correct
    }
    
    if not total_correct:
        total_str = f"{total:.3f}" if total is not None else "0.000"
        resultat["alerte"] = f"Attention: la somme des articles ({somme_articles:.3f} DT) ne correspond pas au total ({total_str} DT)!"
    
    return resultat

def post_process_timbre_fiscal(result_data):
    """
    Post-traitement pour s'assurer que le timbre fiscal est bien détecté et inclus dans les articles
    """
    if not result_data or not isinstance(result_data, dict):
        return result_data
    
    articles = result_data.get("Articles", [])
    
    # Chercher le timbre fiscal dans les articles
    for article in articles:
        if isinstance(article, dict):
            nom = article.get("nom", "").lower()
            prix = article.get("prix", "")
            
            # Détecter le timbre fiscal par le montant (0.100 DT, 0.200 DT, etc.)
            if prix in ["0.100 DT", "0.200 DT", "0.300 DT", "0.400 DT", "0.500 DT"]:
                # S'assurer que le nom est "TIMBRE FISCAL"
                if "timbre" not in nom and "fiscal" not in nom:
                    article["nom"] = "TIMBRE FISCAL"
                    print(f"Nom de l'article timbre fiscal corrigé: {article['nom']}")
                break
            
            # Détecter par le nom aussi
            if any(keyword in nom for keyword in ["timbre", "fiscal", "taxe", "stamp"]):
                # S'assurer que le nom est "TIMBRE FISCAL"
                article["nom"] = "TIMBRE FISCAL"
                print(f"Nom de l'article timbre fiscal corrigé: {article['nom']}")
                break
            
            # Conversion des montants : 100 DT → 0.100 DT pour les timbres fiscaux
            if prix == "100 DT" and any(keyword in nom for keyword in ["timbre", "fiscal", "taxe", "stamp"]):
                article["prix"] = "0.100 DT"
                article["nom"] = "TIMBRE FISCAL"
                print(f"Montant timbre fiscal converti: 100 DT → 0.100 DT")
                break
    
    return result_data

def valider_et_corriger_avec_regex(result_data, texte_ocr):
    """
    Valide et corrige les résultats LLM avec l'extraction regex
    """
    if not result_data or not isinstance(result_data, dict):
        return result_data
    
    # Extraire les éléments avec regex
    regex_result = extraire_elements_avec_regex(texte_ocr)
    
    print("=== VALIDATION REGEX ===")
    print(f"Dates trouvées: {regex_result.get('dates_valides', [])}")
    print(f"Timbres fiscaux: {regex_result.get('timbres_fiscaux', [])}")
    print(f"Total regex: {regex_result.get('total', '')}")
    print(f"Articles regex: {len(regex_result.get('articles', []))}")
    print(f"Total cohérent: {regex_result.get('total_coherent', False)}")
    print("========================")
    
    # Corriger la date si nécessaire
    if not result_data.get("Date") and regex_result.get("dates_valides"):
        result_data["Date"] = regex_result["dates_valides"][0]
        print(f"Date corrigée avec regex: {result_data['Date']}")
    
    
    
    # Corriger le total si nécessaire
    if not result_data.get("Total") and regex_result.get("total"):
        result_data["Total"] = regex_result["total"]
        print(f"Total corrigé avec regex: {result_data['Total']}")
    
    # Corriger les articles si nécessaire (si le LLM n'a pas bien détecté)
    if not result_data.get("Articles") and regex_result.get("articles"):
        result_data["Articles"] = regex_result["articles"]
        print(f"Articles corrigés avec regex: {len(result_data['Articles'])} articles")
    
    # Ajouter des informations de validation
    if regex_result.get("alerte"):
        result_data["AlerteValidation"] = regex_result["alerte"]
        print(f"Alerte ajoutée: {regex_result['alerte']}")
    
    result_data["ValidationRegex"] = {
        "total_coherent": regex_result.get("total_coherent", False),
        "somme_articles": regex_result.get("somme_articles", ""),
        "total_detecte": regex_result.get("total", "")
    }
    
    return result_data

def extract_text_doctr(file_path):
    try:
        print(f"Doctr: Processing {file_path}")
        # Charger le document depuis le chemin
        doc = DocumentFile.from_images(file_path)

        # Charger le modèle Doctr
        model = ocr_predictor(pretrained=True)

        # Faire la prédiction
        result = model(doc)

        # Extraire le texte brut
        extracted_text = result.render()
        print(f"Doctr: Extracted {len(extracted_text)} characters")
        return extracted_text
    except Exception as e:
        error_msg = f"Erreur Doctr: {str(e)}"
        print(error_msg)
        return error_msg

def extract_text_docling(file_path):
    try:
        print(f"Docling: Processing {file_path}")
        converter = DocumentConverter()
        path_obj = Path(file_path)
        result = converter.convert(path_obj)
        document = result.document

        text_lines = {}
        if hasattr(document, 'texts'):
            for text_item in document.texts:
                if hasattr(text_item, 'prov') and text_item.prov:
                    y_pos = sum([prov.bbox.t for prov in text_item.prov]) / len(text_item.prov)
                    text_lines[y_pos] = text_lines.get(y_pos, "") + " " + text_item.text

        sorted_lines = sorted(text_lines.items(), key=lambda x: -x[0])
        extracted_text = "\n".join([line[1].strip() for line in sorted_lines])
        print(f"Docling: Extracted {len(extracted_text)} characters")
        return extracted_text

    except Exception as e:
        error_msg = f"Erreur Docling: {str(e)}"
        print(error_msg)
        return error_msg

def extract_text_tesseract(file_path):
    try:
        print(f"Tesseract: Processing {file_path}")
        
        # Vérifier si Tesseract est disponible
        try:
            import pytesseract
            # Test rapide pour vérifier si Tesseract est installé
            pytesseract.get_tesseract_version()
        except Exception as tesseract_error:
            error_msg = f"Tesseract non disponible: {str(tesseract_error)}\n\nPour installer Tesseract:\n1. Téléchargez: https://github.com/UB-Mannheim/tesseract/wiki\n2. Installez dans C:\\Program Files\\Tesseract-OCR\\\n3. Ajoutez au PATH système"
            print(error_msg)
            return error_msg
        
        ext = os.path.splitext(file_path)[1].lower()
        images = []
        if ext == '.pdf':
            try:
                from pdf2image import convert_from_path
                images = convert_from_path(file_path)
            except ImportError:
                return "Erreur: pdf2image non installé pour traiter les PDF"
        else:
            images = [Image.open(file_path)]
        
        text = ""
        for img in images:
            text += pytesseract.image_to_string(img, lang='fra+eng') + "\n"
        extracted_text = text.strip()
        print(f"Tesseract: Extracted {len(extracted_text)} characters")
        return extracted_text
    except Exception as e:
        error_msg = f"Erreur Tesseract: {str(e)}"
        print(error_msg)
        return error_msg

def save_ticket_to_history(llm_analysis):
    """
    Sauvegarde un ticket analysé dans l'historique
    """
    if not llm_analysis or not isinstance(llm_analysis, dict):
        return None
    
    try:
        # Extraire les données du ticket
        date_str = llm_analysis.get("Date", "")
        magasin = llm_analysis.get("Magasin", "Magasin inconnu")
        total_str = llm_analysis.get("Total", "0.000 DT")
        numero_ticket = llm_analysis.get("NumeroTicket", "")
        articles = llm_analysis.get("Articles", [])
        
        # Nettoyer le total
        try:
            total_clean = total_str.replace(" DT", "").replace(",", ".")
            total_decimal = Decimal(total_clean)
        except:
            total_decimal = Decimal('0.000')
        
        # Parser la date
        try:
            if date_str:
                # Essayer différents formats de date
                date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
                ticket_date = None
                for fmt in date_formats:
                    try:
                        ticket_date = datetime.strptime(date_str.split()[0], fmt).date()
                        break
                    except:
                        continue
                if not ticket_date:
                    ticket_date = date.today()
            else:
                ticket_date = date.today()
        except:
            ticket_date = date.today()
        
        # Créer l'entrée dans l'historique
        ticket_history = TicketHistory.objects.create(
            date_ticket=ticket_date,
            magasin=magasin,
            total=total_decimal,
            numero_ticket=numero_ticket,
            articles_data=articles,
            llm_analysis=llm_analysis
        )
        
        print(f"Ticket sauvegardé: {ticket_history}")
        return ticket_history
        
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du ticket: {e}")
        return None

def generate_accounting_report(llm_analysis, compte="606100", description="Achat divers", save_to_db=True):
    """
    Génère un rapport comptable à partir des données LLM
    """
    if not llm_analysis or not isinstance(llm_analysis, dict):
        return None
    
    # Extraire les données du ticket
    date_ticket = llm_analysis.get("Date", "")
    magasin = llm_analysis.get("Magasin", "Magasin inconnu")
    total = llm_analysis.get("Total", "0.000 DT")
    
    # Nettoyer le total (enlever "DT" et convertir en float)
    try:
        total_clean = total.replace(" DT", "").replace(",", ".")
        total_float = float(total_clean)
    except:
        total_float = 0.0
    
    # Créer le libellé d'écriture
    libelle_ecriture = f"Achat-{magasin}"
    
    # Créer le rapport comptable
    report_data = {
        "date_ticket": date_ticket,
        "compte": compte,
        "description": description,
        "libelle_ecriture": libelle_ecriture,
        "debit": f"{total_float:.3f}",
        "credit": "",
        "magasin": magasin,
        "total_original": total
    }
    
    # Sauvegarder dans la base de données si demandé
    if save_to_db:
        try:
            # Sauvegarder le ticket dans l'historique
            ticket_history = save_ticket_to_history(llm_analysis)
            
            if ticket_history:
                # Créer l'entrée comptable
                try:
                    # Parser la date pour l'écriture comptable
                    date_ecriture = date.today()
                    if date_ticket:
                        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
                        for fmt in date_formats:
                            try:
                                date_ecriture = datetime.strptime(date_ticket.split()[0], fmt).date()
                                break
                            except:
                                continue
                    
                    accounting_entry = AccountingEntry.objects.create(
                        ticket=ticket_history,
                        date_ecriture=date_ecriture,
                        compte=compte,
                        description=description,
                        libelle_ecriture=libelle_ecriture,
                        debit=Decimal(f"{total_float:.3f}"),
                        credit=None
                    )
                    print(f"Entrée comptable créée: {accounting_entry}")
                except Exception as e:
                    print(f"Erreur lors de la création de l'entrée comptable: {e}")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde: {e}")
    
    return report_data

def generate_accounting_pdf(report_data):
    """
    Génère un PDF du bilan comptable
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titre
    title = Paragraph("Bilan Comptable - Ticket de Caisse", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informations du ticket
    ticket_info = [
        ["Date du ticket:", report_data["date_ticket"]],
        ["Magasin:", report_data["magasin"]],
        ["Total:", report_data["total_original"]],
    ]
    
    ticket_table = Table(ticket_info, colWidths=[2*inch, 4*inch])
    ticket_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(ticket_table)
    elements.append(Spacer(1, 20))
    
    # Tableau comptable
    accounting_data = [
        ["Date", "Compte", "Description", "Libellé Écriture", "Débit", "Crédit"],
        [
            report_data["date_ticket"],
            report_data["compte"],
            report_data["description"],
            report_data["libelle_ecriture"],
            f"{report_data['debit']} DT",
            report_data["credit"]
        ]
    ]
    
    accounting_table = Table(accounting_data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
    accounting_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, 1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(accounting_table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_accounting_excel(report_data):
    """
    Génère un fichier Excel du bilan comptable
    """
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan Comptable"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "Bilan Comptable - Ticket de Caisse"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Informations du ticket (ligne 3-5)
    ws['A3'] = "Date du ticket:"
    ws['B3'] = report_data["date_ticket"]
    ws['A4'] = "Magasin:"
    ws['B4'] = report_data["magasin"]
    ws['A5'] = "Total:"
    ws['B5'] = report_data["total_original"]
    
    # Style pour les labels
    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # En-têtes du tableau comptable (ligne 7)
    headers = ["Date", "Compte", "Description", "Libellé Écriture", "Débit", "Crédit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Données comptables (ligne 8)
    data_row = [
        report_data["date_ticket"],
        report_data["compte"],
        report_data["description"],
        report_data["libelle_ecriture"],
        f"{report_data['debit']} DT",
        report_data["credit"]
    ]
    
    for col, value in enumerate(data_row, 1):
        cell = ws.cell(row=8, column=col, value=value)
        cell.border = border
        cell.alignment = center_alignment
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 12, 20, 25, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_cumulative_accounting_pdf(start_date=None, end_date=None):
    """
    Génère un PDF cumulatif de toutes les entrées comptables
    """
    # Récupérer toutes les entrées comptables
    entries = AccountingEntry.objects.all()
    
    if start_date:
        entries = entries.filter(date_ecriture__gte=start_date)
    if end_date:
        entries = entries.filter(date_ecriture__lte=end_date)
    
    entries = entries.order_by('date_ecriture')
    
    if not entries.exists():
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titre
    title = Paragraph("Bilan Comptable Cumulatif", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Informations générales
    total_debit = sum(entry.debit for entry in entries)
    nb_entries = entries.count()
    
    summary_info = [
        ["Période:", f"{start_date or 'Début'} à {end_date or 'Aujourd\'hui'}"],
        ["Nombre d'écritures:", str(nb_entries)],
        ["Total débit:", f"{total_debit:.3f} DT"],
    ]
    
    summary_table = Table(summary_info, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Tableau des écritures comptables
    accounting_data = [
        ["Date", "Compte", "Description", "Libellé Écriture", "Débit", "Crédit"]
    ]
    
    for entry in entries:
        accounting_data.append([
            entry.date_ecriture.strftime("%d/%m/%Y"),
            entry.compte,
            entry.description,
            entry.libelle_ecriture,
            f"{entry.debit:.3f} DT",
            f"{entry.credit:.3f} DT" if entry.credit else ""
        ])
    
    # Ajouter une ligne de total
    accounting_data.append([
        "", "", "", "TOTAL", f"{total_debit:.3f} DT", ""
    ])
    
    accounting_table = Table(accounting_data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1*inch, 1*inch])
    accounting_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(accounting_table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_cumulative_accounting_excel(start_date=None, end_date=None):
    """
    Génère un fichier Excel cumulatif de toutes les entrées comptables
    """
    # Récupérer toutes les entrées comptables
    entries = AccountingEntry.objects.all()
    
    if start_date:
        entries = entries.filter(date_ecriture__gte=start_date)
    if end_date:
        entries = entries.filter(date_ecriture__lte=end_date)
    
    entries = entries.order_by('date_ecriture')
    
    if not entries.exists():
        return None
    
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan Comptable Cumulatif"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "Bilan Comptable Cumulatif"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Informations générales
    total_debit = sum(entry.debit for entry in entries)
    nb_entries = entries.count()
    
    ws['A3'] = "Période:"
    ws['B3'] = f"{start_date or 'Début'} à {end_date or 'Aujourd\'hui'}"
    ws['A4'] = "Nombre d'écritures:"
    ws['B4'] = str(nb_entries)
    ws['A5'] = "Total débit:"
    ws['B5'] = f"{total_debit:.3f} DT"
    
    # Style pour les labels
    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # En-têtes du tableau comptable (ligne 7)
    headers = ["Date", "Compte", "Description", "Libellé Écriture", "Débit", "Crédit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Données comptables
    current_row = 8
    for entry in entries:
        data_row = [
            entry.date_ecriture.strftime("%d/%m/%Y"),
            entry.compte,
            entry.description,
            entry.libelle_ecriture,
            f"{entry.debit:.3f} DT",
            f"{entry.credit:.3f} DT" if entry.credit else ""
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            cell.alignment = center_alignment
        
        current_row += 1
    
    # Ligne de total
    total_row = ["", "", "", "TOTAL", f"{total_debit:.3f} DT", ""]
    for col, value in enumerate(total_row, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.border = border
        cell.alignment = center_alignment
        cell.fill = total_fill
        cell.font = Font(bold=True)
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 12, 20, 25, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def download_cumulative_excel(request):
    """
    Vue pour télécharger le bilan comptable cumulatif en Excel
    """
    if request.method == 'POST':
        try:
            # Récupérer les dates de filtrage
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            
            start_date = None
            end_date = None
            
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # Générer le fichier Excel cumulatif
            excel_buffer = generate_cumulative_accounting_excel(start_date, end_date)
            
            if not excel_buffer:
                return HttpResponse("Aucune donnée disponible pour la période sélectionnée", status=400)
            
            # Créer la réponse HTTP
            response = HttpResponse(
                excel_buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bilan_comptable_cumulatif_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as e:
            print(f"Erreur lors de la génération du fichier Excel cumulatif: {e}")
            return HttpResponse(f"Erreur lors de la génération du fichier Excel: {str(e)}", status=500)
    
    return HttpResponse("Méthode non autorisée", status=405)

def view_history(request):
    """
    Vue pour afficher l'historique des tickets avec gestion de budget
    """
    from datetime import datetime
    from .models import Budget
    
    tickets = TicketHistory.objects.all().order_by('-created_at')
    total_amount = sum(ticket.total for ticket in tickets)
    
    # Gestion du budget
    budget_type = request.GET.get('budget_type', 'monthly')
    current_date = datetime.now()
    
    # Récupérer les paramètres de mois et année
    selected_year = int(request.GET.get('year', current_date.year))
    selected_month = int(request.GET.get('month', current_date.month))
    
    # Calculer les dépenses selon le type de budget
    if budget_type == 'monthly':
        # Dépenses du mois sélectionné
        monthly_tickets = tickets.filter(
            date_ticket__year=selected_year,
            date_ticket__month=selected_month
        )
        current_expenses = sum(ticket.total for ticket in monthly_tickets)
        # Récupérer le budget pour le mois/année spécifique
        try:
            budget = Budget.objects.get(
                type_budget='monthly',
                annee=selected_year,
                mois=selected_month
            )
        except Budget.DoesNotExist:
            budget = None
    else:  # yearly
        # Dépenses de l'année sélectionnée
        yearly_tickets = tickets.filter(date_ticket__year=selected_year)
        current_expenses = sum(ticket.total for ticket in yearly_tickets)
        # Récupérer le budget pour l'année spécifique
        try:
            budget = Budget.objects.get(
                type_budget='yearly',
                annee=selected_year
            )
        except Budget.DoesNotExist:
            budget = None
    
    # Calculer les indicateurs de budget
    budget_info = {
        'budget_type': budget_type,
        'budget_amount': budget.montant if budget else None,
        'current_expenses': current_expenses,
        'remaining': None,
        'exceeded': None,
        'budget_exists': budget is not None,
        'current_year': selected_year,
        'current_month': selected_month,
        'selected_year': selected_year,
        'selected_month': selected_month
    }
    
    if budget:
        if current_expenses <= budget.montant:
            budget_info['remaining'] = budget.montant - current_expenses
        else:
            budget_info['exceeded'] = current_expenses - budget.montant
    
    # Identifier les tickets qui causent un dépassement - VERSION OPTIMISÉE
    tickets_with_budget_status = []
    running_total = 0
    budget_exceeded = False
    overflow_threshold_reached = False
    
    if budget:
        # Filtrer et trier les tickets selon la période sélectionnée
        if budget_type == 'monthly':
            period_tickets = tickets.filter(
                date_ticket__year=selected_year,
                date_ticket__month=selected_month
            )
        else:  # yearly
            period_tickets = tickets.filter(
                date_ticket__year=selected_year
            )
        
        # Trier par date et heure de création pour un ordre chronologique précis
        sorted_tickets = period_tickets.order_by('date_ticket', 'created_at')
        
        for ticket in sorted_tickets:
            previous_total = running_total
            running_total += ticket.total
            
            # Déterminer le statut du ticket
            ticket_status = {
                'ticket': ticket,
                'running_total': running_total,
                'causes_overflow': False,
                'is_over_budget': False,
                'percentage_of_budget': (running_total / budget.montant) * 100
            }
            
            # Marquer si ce ticket spécifique cause le dépassement
            if previous_total <= budget.montant and running_total > budget.montant:
                ticket_status['causes_overflow'] = True
                overflow_threshold_reached = True
            
            # Marquer tous les tickets ajoutés après le dépassement
            if overflow_threshold_reached:
                ticket_status['is_over_budget'] = True
            
            tickets_with_budget_status.append(ticket_status)
        
        # Marquer si le budget global est dépassé
        budget_exceeded = running_total > budget.montant
    
    # Si aucun budget n'existe, afficher tous les tickets de la période sélectionnée
    if not budget:
        if budget_type == 'monthly':
            period_tickets = tickets.filter(
                date_ticket__year=selected_year,
                date_ticket__month=selected_month
            ).order_by('-date_ticket', '-created_at')
        else:  # yearly
            period_tickets = tickets.filter(
                date_ticket__year=selected_year
            ).order_by('-date_ticket', '-created_at')
        
        for ticket in period_tickets:
            tickets_with_budget_status.append({
                'ticket': ticket,
                'running_total': None,
                'causes_overflow': False,
                'is_over_budget': False,
                'percentage_of_budget': 0
            })
    
    # Calculer les tickets de la période pour les statistiques
    if budget_type == 'monthly':
        period_tickets_for_stats = tickets.filter(
            date_ticket__year=selected_year,
            date_ticket__month=selected_month
        )
    else:  # yearly
        period_tickets_for_stats = tickets.filter(
            date_ticket__year=selected_year
        )
    
    period_total_amount = sum(ticket.total for ticket in period_tickets_for_stats)
    
    context = {
        'tickets': period_tickets_for_stats,  # Tickets de la période sélectionnée
        'total_amount': total_amount,  # Total global pour les statistiques générales
        'period_total_amount': period_total_amount,  # Total de la période
        'ticket_count': tickets.count(),  # Nombre total global
        'period_ticket_count': period_tickets_for_stats.count(),  # Nombre pour la période
        'budget_info': budget_info,
        'tickets_with_budget_status': tickets_with_budget_status,
        'current_budget_type': budget_type,
        'budget_exceeded': budget_exceeded,
        'overflow_threshold_reached': overflow_threshold_reached
    }
    
    return render(request, 'ocrapp/history.html', context)

@csrf_exempt
def manage_budget(request):
    """
    Vue pour gérer les budgets (création/modification)
    """
    from datetime import datetime
    from .models import Budget
    
    if request.method == 'POST':
        try:
            budget_type = request.POST.get('budget_type')
            montant = float(request.POST.get('montant', 0))
            current_date = datetime.now()
            
            # Récupérer les paramètres de mois et année ou utiliser les valeurs actuelles
            selected_year = int(request.POST.get('year', current_date.year))
            selected_month = int(request.POST.get('month', current_date.month))
            
            if budget_type == 'monthly':
                # Créer ou mettre à jour le budget mensuel pour la période sélectionnée
                budget, created = Budget.objects.get_or_create(
                    type_budget='monthly',
                    annee=selected_year,
                    mois=selected_month,
                    defaults={'montant': montant}
                )
                if not created:
                    budget.montant = montant
                    budget.save()
                    
                message = f"Budget mensuel {selected_month:02d}/{selected_year} {'créé' if created else 'mis à jour'}: {montant} DT"
                
            else:  # yearly
                # Créer ou mettre à jour le budget annuel pour l'année sélectionnée
                budget, created = Budget.objects.get_or_create(
                    type_budget='yearly',
                    annee=selected_year,
                    defaults={'montant': montant}
                )
                if not created:
                    budget.montant = montant
                    budget.save()
                    
                message = f"Budget annuel {selected_year} {'créé' if created else 'mis à jour'}: {montant} DT"
            
            return JsonResponse({
                'success': True,
                'message': message,
                'budget_amount': float(budget.montant),
                'budget_type': budget_type
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la gestion du budget: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

def download_accounting_excel(request):
    """
    Vue pour télécharger le bilan comptable en Excel
    """
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            compte = request.POST.get('compte', '606100')
            description = request.POST.get('description', 'Achat divers')
            analysis_type = request.POST.get('analysis_data', 'qwen')
            
            # Récupérer les données selon le type d'analyse
            if analysis_type == 'gemini':
                analysis_data = request.session.get('gemini_analysis')
                print(f"Utilisation des données Gemini: {list(analysis_data.keys()) if analysis_data else 'Aucune donnée'}")
            else:
                analysis_data = request.session.get('llm_analysis')
                print(f"Utilisation des données Qwen: {list(analysis_data.keys()) if analysis_data else 'Aucune donnée'}")
            
            if not analysis_data:
                print(f"Aucune donnée {analysis_type} trouvée en session")
                return HttpResponse("Aucune donnée de ticket disponible", status=400)
            
            # Générer le rapport comptable et sauvegarder en base
            report_data = generate_accounting_report(analysis_data, compte, description, save_to_db=True)
            if not report_data:
                return HttpResponse("Erreur lors de la génération du rapport", status=400)
            
            # Générer le fichier Excel
            excel_buffer = generate_accounting_excel(report_data)
            
            # Créer la réponse HTTP
            response = HttpResponse(
                excel_buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bilan_comptable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Exception as e:
            print(f"Erreur lors de la génération du fichier Excel: {e}")
            return HttpResponse(f"Erreur lors de la génération du fichier Excel: {str(e)}", status=500)
    
    return HttpResponse("Méthode non autorisée", status=405)

def upload_ticket(request):
    ocr_results = None
    llm_analysis = None
    gemini_analysis = None
    regex_analysis = None
    error = None
    form = TicketUploadForm()
    
    # Run system diagnostics on first load
    if request.method == 'GET':
        system_issues = diagnose_system()
        if system_issues:
            logger.warning("System issues detected - some features may not work properly")
    
    if request.method == 'POST':
        # Debug: afficher les données POST
        print("POST data:", request.POST)
        logger.info("Processing OCR request")
        
        # Vérifier si on a des textes OCR dans les champs cachés
        ocr_doctr = request.POST.get('ocr_doctr')
        ocr_tesseract = request.POST.get('ocr_tesseract')
        ocr_docling = request.POST.get('ocr_docling')
        
        print("OCR fields:", {
            'doctr': bool(ocr_doctr),
            'tesseract': bool(ocr_tesseract),
            'docling': bool(ocr_docling)
        })
        
        if ocr_doctr and ocr_tesseract and ocr_docling:
            print("Using existing OCR texts")
            print(f"OCR data lengths: doctr={len(ocr_doctr)}, tesseract={len(ocr_tesseract)}, docling={len(ocr_docling)}")
            # Cas où on a déjà les textes OCR et on veut analyser avec LLM
            ocr_results = {
                'doctr': ocr_doctr,
                'tesseract': ocr_tesseract,
                'docling': ocr_docling,
            }
            
            # Récupérer l'instance de l'image depuis la session ou la dernière entrée
            try:
                # Essayer de récupérer l'ID de l'image depuis la session
                image_id = request.session.get('current_image_id')
                if image_id:
                    from .models import ExtractionHistory
                    instance = ExtractionHistory.objects.get(id=image_id)
                    form = TicketUploadForm(instance=instance)
                    print(f"Image récupérée depuis la session: {instance.image.url}")
                else:
                    # Fallback: récupérer la dernière image uploadée
                    from .models import ExtractionHistory
                    instance = ExtractionHistory.objects.latest('uploaded_at')
                    form = TicketUploadForm(instance=instance)
                    print(f"Image récupérée (dernière): {instance.image.url}")
            except Exception as e:
                print(f"Erreur lors de la récupération de l'image: {e}")
                # Garder le form vide si on ne peut pas récupérer l'image
                pass
            
            # Vérifier quel type d'analyse est demandé
            analyze_with_gemini = request.POST.get('analyze_gemini') == '1'
            
            if analyze_with_gemini:
                print("Starting Gemini analysis...")
                gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                print(f"Gemini analysis result: {type(gemini_analysis)}")
            else:
                # Analyser avec le LLM par défaut
                print("Starting LLM analysis...")
                llm_analysis = analyze_three_texts_with_llm(ocr_results)
                print(f"LLM analysis result: {type(llm_analysis)}")
            
        else:
            print("Processing new image upload")
            # Cas normal avec upload d'image
            form = TicketUploadForm(request.POST, request.FILES)
            if form.is_valid():
                instance = form.save()
                image_path = instance.image.path
                
                # Sauvegarder l'ID de l'image en session pour les analyses ultérieures
                request.session['current_image_id'] = instance.id
                print(f"Image ID sauvegardé en session: {instance.id}")

                # Vérifier quel bouton a été cliqué
                analyze_all = request.POST.get('ocr_all') == '1'
                analyze_with_llm = request.POST.get('analyze_llm') == '1'
                analyze_with_gemini = request.POST.get('analyze_gemini') == '1'
                analyze_with_regex = request.POST.get('analyze_regex') == '1'
                
                print("Analysis flags:", {
                    'analyze_all': analyze_all,
                    'analyze_with_llm': analyze_with_llm,
                    'analyze_with_gemini': analyze_with_gemini,
                    'analyze_with_regex': analyze_with_regex
                })
                
                # Si on clique sur "Extraction OCR avec Doctr" (bouton bleu)
                if analyze_all and not analyze_with_llm:
                    print("Extracting OCR texts only")
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                
                # Si on clique sur "OCR + Analyse Qwen3-30B" (bouton vert)
                elif analyze_with_llm and analyze_all and not analyze_with_gemini:
                    print("Extracting OCR and analyzing with LLM")
                    # Extraire les OCR puis analyser avec le LLM
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec le LLM
                    if ocr_results:
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)
                
                # Si on clique sur "OCR + Analyse Google Gemini" (bouton violet)
                elif analyze_with_gemini and analyze_all and not analyze_with_llm:
                    print("Extracting OCR and analyzing with Gemini")
                    # Extraire les OCR puis analyser avec Gemini
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec Gemini
                    if ocr_results:
                        gemini_analysis = analyze_three_texts_with_gemini(ocr_results)
                        # Ajouter aussi l'analyse regex pour comparaison
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)
                
                # Si on clique sur "Analyse Regex (Rapide)" (bouton jaune)
                elif analyze_with_regex and not analyze_all:
                    print("Extracting OCR and analyzing with Regex")
                    # Extraire les OCR puis analyser avec regex
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec regex
                    if ocr_results:
                        texte_combine = f"{ocr_results['doctr']}\n{ocr_results['tesseract']}\n{ocr_results['docling']}"
                        regex_analysis = extraire_elements_avec_regex(texte_combine)
                        print(f"Regex analysis result: {regex_analysis}")
                
                # Si on clique sur "Analyse Complète" (bouton info) - les deux flags sont à 1
                elif analyze_all and analyze_with_llm:
                    print("Complete analysis: OCR + LLM")
                    # Lancer les 3 OCR
                    ocr_results = {
                        'tesseract': extract_text_tesseract(image_path),
                        'doctr': extract_text_doctr(image_path),
                        'docling': extract_text_docling(image_path),
                    }
                    
                    # Puis analyser avec le LLM
                    if ocr_results:
                        llm_analysis = analyze_three_texts_with_llm(ocr_results)

    # Convertir les données LLM en JSON pour le formulaire
    llm_analysis_json = None
    if llm_analysis and not isinstance(llm_analysis, str):
        try:
            # Filtrer les données sensibles et s'assurer que les valeurs sont sérialisables
            clean_llm_data = {}
            for key, value in llm_analysis.items():
                if key not in ['texte_fusionne', 'raw_response']:
                    # S'assurer que la valeur est sérialisable
                    if isinstance(value, (str, int, float, list, dict, bool)) or value is None:
                        clean_llm_data[key] = value
            
            # Sauvegarder en session pour le téléchargement PDF
            request.session['llm_analysis'] = clean_llm_data
            
            llm_analysis_json = json.dumps(clean_llm_data, ensure_ascii=False, separators=(',', ':'))
            print(f"JSON généré: {llm_analysis_json[:200]}...")
        except Exception as e:
            print(f"Erreur lors de la génération du JSON: {e}")
            llm_analysis_json = None
    
    # Sauvegarder gemini_analysis en session si disponible
    if gemini_analysis and not isinstance(gemini_analysis, str):
        try:
            # Filtrer les données sensibles et s'assurer que les valeurs sont sérialisables
            clean_gemini_data = {}
            for key, value in gemini_analysis.items():
                if key not in ['texte_fusionne', 'raw_response']:
                    # S'assurer que la valeur est sérialisable
                    if isinstance(value, (str, int, float, list, dict, bool)) or value is None:
                        clean_gemini_data[key] = value
            
            # Sauvegarder en session pour le téléchargement PDF
            request.session['gemini_analysis'] = clean_gemini_data
            print(f"Données Gemini sauvegardées en session: {list(clean_gemini_data.keys())}")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde Gemini en session: {e}")
    
    return render(request, 'ocrapp/upload_ticket.html', {
        'form': form,
        'ocr_results': ocr_results,
        'llm_analysis': llm_analysis,
        'gemini_analysis': gemini_analysis,
        'regex_analysis': regex_analysis,
        'llm_analysis_json': llm_analysis_json,
        'error': error
    })

@csrf_exempt
def filter_accounting_data(request):
    """
    Filtre les données comptables par date pour l'affichage AJAX
    """
    if request.method == 'POST':
        try:
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            # Construire la requête de base
            entries_query = AccountingEntry.objects.all()
            
            # Appliquer les filtres de date si fournis
            if start_date:
                try:
                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                    entries_query = entries_query.filter(date_ecriture__gte=start_date_obj)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Format de date de début invalide'})
            
            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                    entries_query = entries_query.filter(date_ecriture__lte=end_date_obj)
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Format de date de fin invalide'})
            
            # Récupérer les entrées filtrées
            entries = entries_query.order_by('-date_ecriture')
            
            # Préparer les données pour JSON
            entries_data = []
            for entry in entries:
                entries_data.append({
                    'date_ecriture': entry.date_ecriture.strftime('%Y-%m-%d'),
                    'compte': entry.compte,
                    'description': entry.description,
                    'libelle_ecriture': entry.libelle_ecriture,
                    'debit': float(entry.debit),
                    'credit': float(entry.credit) if entry.credit else 0.0
                })
            
            # Calculer les totaux par compte
            account_totals = {}
            total_debit = Decimal('0')
            total_credit = Decimal('0')
            
            for entry in entries:
                account = entry.compte
                if account not in account_totals:
                    account_totals[account] = {'debit': Decimal('0'), 'credit': Decimal('0')}
                
                account_totals[account]['debit'] += entry.debit
                if entry.credit:
                    account_totals[account]['credit'] += entry.credit
                
                total_debit += entry.debit
                if entry.credit:
                    total_credit += entry.credit
            
            # Convertir les totaux en float pour JSON
            account_totals_json = {}
            for account, totals in account_totals.items():
                account_totals_json[account] = {
                    'debit': float(totals['debit']),
                    'credit': float(totals['credit'])
                }
            
            # Résumé général
            summary = {
                'total_entries': len(entries_data),
                'total_debit': float(total_debit),
                'total_credit': float(total_credit),
                'total_accounts': len(account_totals),
                'period': {
                    'start': start_date if start_date else 'Début',
                    'end': end_date if end_date else 'Fin'
                }
            }
            
            return JsonResponse({
                'success': True,
                'entries': entries_data,
                'account_totals': account_totals_json,
                'summary': summary
            })
            
        except Exception as e:
            logger.error(f"Erreur lors du filtrage des données comptables: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Erreur serveur: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


import google.generativeai as genai
genai.configure(api_key=os.environ.get("GOOGLE_API_KEY"))
def verifier_par_gemini(qwen_data, ocr_texts_combined):
    """
    Utilise Gemini 2.5 Flash pour valider les informations extraites par Qwen
    """
    prompt = f"""
Tu es un assistant spécialisé dans la validation de données extraites de tickets de caisse.

Voici un texte brut issu de plusieurs OCR :
-------------------------------
{ocr_texts_combined}
-------------------------------

Et voici les données extraites par un autre modèle (Qwen) :

{json.dumps(qwen_data, indent=2, ensure_ascii=False)}

Analyse attentivement le texte OCR et dis-moi si les informations suivantes semblent cohérentes avec le contenu :
- Le nom du magasin
- La date et heure
- Le numéro du ticket
- Les articles (noms et prix)
- Le montant total

RÉPONDS uniquement avec un JSON structuré comme ceci :

{{
  "verdict": "valide" ou "invalide",
  "problemes_detectes": ["description 1", "description 2", ...],
  "suggestions": ["correction 1", "correction 2", ...]
}}

IMPORTANT : ne donne que ce JSON.
    """

    try:
        model = genai.GenerativeModel("gemini-1.5-flash")  # Ou "gemini-2.5-flash" si dispo
        response = model.generate_content(prompt)
        return clean_json_response(response.text)
    except Exception as e:
        print("Erreur avec Gemini:", str(e))
        return {
            "verdict": "erreur",
            "problemes_detectes": [f"Erreur Gemini: {str(e)}"],
            "suggestions": []
        }

@csrf_exempt
def get_ticket_details(request, ticket_id):
    """
    Vue pour récupérer les détails d'un ticket pour modification
    """
    if request.method == 'GET':
        try:
            ticket = Ticket.objects.get(id=ticket_id)
            return JsonResponse({
                'success': True,
                'ticket': {
                    'id': ticket.id,
                    'magasin': ticket.magasin,
                    'numero_ticket': ticket.numero_ticket,
                    'date_ticket': ticket.date_ticket.strftime('%Y-%m-%d'),
                    'total': float(ticket.total),
                    'articles_data': ticket.articles_data or []
                }
            })
        except Ticket.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Ticket non trouvé'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la récupération du ticket: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@csrf_exempt
def update_ticket(request, ticket_id):
    """
    Vue pour mettre à jour un ticket
    """
    if request.method == 'POST':
        try:
            ticket = Ticket.objects.get(id=ticket_id)
            
            # Récupérer les données du formulaire
            magasin = request.POST.get('magasin')
            numero_ticket = request.POST.get('numero_ticket')
            date_ticket = request.POST.get('date_ticket')
            total = request.POST.get('total')
            articles_data = request.POST.get('articles_data')
            
            # Mettre à jour les champs
            if magasin:
                ticket.magasin = magasin
            if numero_ticket:
                ticket.numero_ticket = numero_ticket
            if date_ticket:
                ticket.date_ticket = datetime.strptime(date_ticket, '%Y-%m-%d').date()
            if total:
                ticket.total = Decimal(total)
            if articles_data:
                import json
                ticket.articles_data = json.loads(articles_data)
            
            ticket.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket mis à jour avec succès',
                'ticket': {
                    'id': ticket.id,
                    'magasin': ticket.magasin,
                    'numero_ticket': ticket.numero_ticket,
                    'date_ticket': ticket.date_ticket.strftime('%Y-%m-%d'),
                    'total': float(ticket.total),
                    'articles_data': ticket.articles_data or []
                }
            })
            
        except Ticket.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Ticket non trouvé'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la mise à jour du ticket: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
@ c s r f _ e x e m p t  
 d e f   s a v e _ t i c k e t _ a n a l y s i s ( r e q u e s t ) :  
         " " "  
         V u e   p o u r   s a u v e g a r d e r   l e s   d o n n � � e s   m o d i f i � � e s   d ' u n   t i c k e t   a p r � � s   a n a l y s e   L L M  
         " " "  
         i f   r e q u e s t . m e t h o d   = =   ' P O S T ' :  
                 t r y :  
                         #   R � � c u p � � r e r   l e s   d o n n � � e s   d u   f o r m u l a i r e  
                         m a g a s i n   =   r e q u e s t . P O S T . g e t ( ' m a g a s i n ' )  
                         d a t e _ s t r   =   r e q u e s t . P O S T . g e t ( ' d a t e ' )  
                         n u m e r o _ t i c k e t   =   r e q u e s t . P O S T . g e t ( ' n u m e r o ' )  
                         t o t a l _ s t r   =   r e q u e s t . P O S T . g e t ( ' t o t a l ' )  
                         t v a _ r a t e _ s t r   =   r e q u e s t . P O S T . g e t ( ' t v a _ r a t e ' )  
                         t v a _ a m o u n t _ s t r   =   r e q u e s t . P O S T . g e t ( ' t v a _ a m o u n t ' )  
                         a r t i c l e s _ j s o n   =   r e q u e s t . P O S T . g e t ( ' a r t i c l e s ' )  
                         a n a l y s i s _ t y p e   =   r e q u e s t . P O S T . g e t ( ' a n a l y s i s _ t y p e ' ,   ' q w e n ' )     #   q w e n   o u   g e m i n i  
                          
                         #   V a l i d a t i o n   d e s   d o n n � � e s   o b l i g a t o i r e s  
                         i f   n o t   a l l ( [ m a g a s i n ,   d a t e _ s t r ,   t o t a l _ s t r ] ) :  
                                 r e t u r n   J s o n R e s p o n s e ( {  
                                         ' s u c c e s s ' :   F a l s e ,  
                                         ' e r r o r ' :   ' L e s   c h a m p s   m a g a s i n ,   d a t e   e t   t o t a l   s o n t   o b l i g a t o i r e s '  
                                 } )  
                          
                         #   C o n v e r s i o n   d e s   t y p e s  
                         t r y :  
                                 d a t e _ t i c k e t   =   d a t e t i m e . s t r p t i m e ( d a t e _ s t r ,   ' % Y - % m - % d ' ) . d a t e ( )  
                                 t o t a l   =   D e c i m a l ( t o t a l _ s t r )  
                                 t v a _ r a t e   =   D e c i m a l ( t v a _ r a t e _ s t r )   i f   t v a _ r a t e _ s t r   e l s e   D e c i m a l ( ' 0 ' )  
                                 t v a _ a m o u n t   =   D e c i m a l ( t v a _ a m o u n t _ s t r )   i f   t v a _ a m o u n t _ s t r   e l s e   D e c i m a l ( ' 0 ' )  
                         e x c e p t   ( V a l u e E r r o r ,   I n v a l i d O p e r a t i o n )   a s   e :  
                                 r e t u r n   J s o n R e s p o n s e ( {  
                                         ' s u c c e s s ' :   F a l s e ,  
                                         ' e r r o r ' :   f ' E r r e u r   d e   f o r m a t   d a n s   l e s   d o n n � � e s :   { s t r ( e ) } '  
                                 } )  
                          
                         #   P a r s e r   l e s   a r t i c l e s  
                         a r t i c l e s _ d a t a   =   [ ]  
                         i f   a r t i c l e s _ j s o n :  
                                 t r y :  
                                         i m p o r t   j s o n  
                                         a r t i c l e s _ d a t a   =   j s o n . l o a d s ( a r t i c l e s _ j s o n )  
                                 e x c e p t   j s o n . J S O N D e c o d e E r r o r :  
                                         r e t u r n   J s o n R e s p o n s e ( {  
                                                 ' s u c c e s s ' :   F a l s e ,  
                                                 ' e r r o r ' :   ' F o r m a t   J S O N   i n v a l i d e   p o u r   l e s   a r t i c l e s '  
                                         } )  
                          
                         #   C r � � e r   o u   m e t t r e   � �   j o u r   l e   t i c k e t  
                         t i c k e t ,   c r e a t e d   =   T i c k e t . o b j e c t s . g e t _ o r _ c r e a t e (  
                                 m a g a s i n = m a g a s i n ,  
                                 d a t e _ t i c k e t = d a t e _ t i c k e t ,  
                                 n u m e r o _ t i c k e t = n u m e r o _ t i c k e t   o r   ' ' ,  
                                 d e f a u l t s = {  
                                         ' t o t a l ' :   t o t a l ,  
                                         ' a r t i c l e s _ d a t a ' :   a r t i c l e s _ d a t a ,  
                                         ' t v a _ r a t e ' :   t v a _ r a t e ,  
                                         ' t v a _ a m o u n t ' :   t v a _ a m o u n t  
                                 }  
                         )  
                          
                         i f   n o t   c r e a t e d :  
                                 #   M e t t r e   � �   j o u r   l e   t i c k e t   e x i s t a n t  
                                 t i c k e t . t o t a l   =   t o t a l  
                                 t i c k e t . a r t i c l e s _ d a t a   =   a r t i c l e s _ d a t a  
                                 t i c k e t . t v a _ r a t e   =   t v a _ r a t e  
                                 t i c k e t . t v a _ a m o u n t   =   t v a _ a m o u n t  
                                 t i c k e t . s a v e ( )  
                          
                         #   C r � � e r   l e s   e n t r � � e s   c o m p t a b l e s  
                         c o m p t e   =   r e q u e s t . P O S T . g e t ( ' c o m p t e ' ,   ' 6 0 6 1 0 0 ' )  
                         d e s c r i p t i o n   =   r e q u e s t . P O S T . g e t ( ' d e s c r i p t i o n ' ,   ' A c h a t   d i v e r s ' )  
                          
                         #   S u p p r i m e r   l e s   a n c i e n n e s   e n t r � � e s   c o m p t a b l e s   p o u r   c e   t i c k e t  
                         A c c o u n t i n g E n t r y . o b j e c t s . f i l t e r (  
                                 n u m e r o _ c o m p t e = c o m p t e ,  
                                 d a t e _ e c r i t u r e = d a t e _ t i c k e t ,  
                                 l i b e l l e _ e c r i t u r e _ _ i c o n t a i n s = m a g a s i n  
                         ) . d e l e t e ( )  
                          
                         #   C r � � e r   l e s   n o u v e l l e s   e n t r � � e s   c o m p t a b l e s  
                         #   E n t r � � e   a u   d � � b i t  
                         A c c o u n t i n g E n t r y . o b j e c t s . c r e a t e (  
                                 d a t e _ e c r i t u r e = d a t e _ t i c k e t ,  
                                 n u m e r o _ c o m p t e = c o m p t e ,  
                                 d e s c r i p t i o n = d e s c r i p t i o n ,  
                                 l i b e l l e _ e c r i t u r e = f ' { m a g a s i n }   -   { n u m e r o _ t i c k e t   o r   " T i c k e t " } ' ,  
                                 d e b i t = t o t a l ,  
                                 c r e d i t = D e c i m a l ( ' 0 ' )  
                         )  
                          
                         #   E n t r � � e   a u   c r � � d i t   ( c o m p t e   d e   t r � � s o r e r i e )  
                         A c c o u n t i n g E n t r y . o b j e c t s . c r e a t e (  
                                 d a t e _ e c r i t u r e = d a t e _ t i c k e t ,  
                                 n u m e r o _ c o m p t e = ' 5 3 1 2 0 0 ' ,     #   C o m p t e   b a n q u e  
                                 d e s c r i p t i o n = ' P a i e m e n t   f o u r n i s s e u r ' ,  
                                 l i b e l l e _ e c r i t u r e = f ' P a i e m e n t   { m a g a s i n }   -   { n u m e r o _ t i c k e t   o r   " T i c k e t " } ' ,  
                                 d e b i t = D e c i m a l ( ' 0 ' ) ,  
                                 c r e d i t = t o t a l  
                         )  
                          
                         r e t u r n   J s o n R e s p o n s e ( {  
                                 ' s u c c e s s ' :   T r u e ,  
                                 ' m e s s a g e ' :   f ' T i c k e t   { " c r � � � � "   i f   c r e a t e d   e l s e   " m i s   � �   j o u r " }   a v e c   s u c c � � s ' ,  
                                 ' t i c k e t _ i d ' :   t i c k e t . i d ,  
                                 ' c r e a t e d ' :   c r e a t e d  
                         } )  
                          
                 e x c e p t   E x c e p t i o n   a s   e :  
                         r e t u r n   J s o n R e s p o n s e ( {  
                                 ' s u c c e s s ' :   F a l s e ,  
                                 ' e r r o r ' :   f ' E r r e u r   l o r s   d e   l a   s a u v e g a r d e :   { s t r ( e ) } '  
                         } )  
          
         r e t u r n   J s o n R e s p o n s e ( { ' s u c c e s s ' :   F a l s e ,   ' e r r o r ' :   ' M � � t h o d e   n o n   a u t o r i s � � e ' } )  
 